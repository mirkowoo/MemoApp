import tkinter as tk
import os
import subprocess
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from search_window import search_data
from openpyxl.utils.dataframe import dataframe_to_rows


#borra el texto de los campos
def clear_entries(entry_fields):
    for entry in entry_fields.values():
        entry.delete(0, tk.END)
#actualiza la tabla de abajo del formulario
def update_table(table_text, editable=False):
    table_text.config(state=tk.NORMAL)
    table_text.delete(1.0, tk.END)

    # Read the Excel file and display the table
    try:
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        col_widths = [max(len(str(cell)) for cell in col) for col in zip(*sheet.iter_rows(values_only=True))]

        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_text = f'{row_number}. '
            for cell, width in zip(row, col_widths):
                row_text += f'{str(cell):<{width + 2}}'
            table_text.insert(tk.END, row_text + '\n')

    except FileNotFoundError:
        table_text.insert(tk.END, "No data available.\n")

    if not editable:
        table_text.config(state=tk.DISABLED)


def get_selected_rows(table_text):
    selected_rows = table_text.tag_ranges(tk.SEL)
    return selected_rows

# def open_excel_file():
#     # Ask the user to select an Excel file
#     file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

#     if file_path:
#         # Get the directory of the selected Excel file
#         file_directory = os.path.dirname(file_path)

#         # Open the file explorer in the directory
#         try:
#             # On Windows
#             subprocess.Popen(['explorer', file_path], shell=True)
#         except FileNotFoundError:
#             # On Linux or macOS (replace 'explorer' with the appropriate file manager command)
#             subprocess.Popen(['xdg-open', file_directory])

#         # Optionally, you can print the selected Excel file path
#         print(f'Selected Excel file: {file_path}')

def delete_selected_rows(table_text, row_number_entry, row_display_label):
    try:
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
        row_to_delete = int(row_number_entry.get())

        # Check if the row_to_delete is within the valid range
        if 2 <= row_to_delete <= sheet.max_row:
            # Delete the row using openpyxl's delete_rows function
            sheet.delete_rows(row_to_delete)
            
            # Save the changes to the workbook
            workbook.save('data.xlsx')
            
            # Update the label to indicate the deletion
            row_display_label.config(text=f'Fila eliminada: {row_to_delete}')
            
            # Update the table display
            update_table(table_text)
        else:
            row_display_label.config(text='Numero inválido de fila. Por favor ingresa un numero entero.')

    except ValueError:
        row_display_label.config(text='Numero inválido de fila. Por favor ingresa un numero entero.')
    except FileNotFoundError:
        row_display_label.config(text="Error: Archivo no encontrado.")


def create_search_window(root):
    search_window = tk.Toplevel(root)
    search_window.title("Search Window")

    tipo_solicitud_entry = tk.Entry(search_window)
    tipo_solicitud_entry.grid(row=0, column=1, padx=5, pady=5)

    search_button = tk.Button(search_window, text="Buscar", command=lambda: search_data(root, tipo_solicitud_entry, result_label))
    search_button.grid(row=0, column=2, padx=5, pady=5)

    result_label = tk.Label(search_window, text="")
    result_label.grid(row=1, column=0, columnspan=3, padx=5, pady=5)
