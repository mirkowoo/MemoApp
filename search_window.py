import tkinter as tk
from openpyxl import load_workbook

def search_data(root, username_entry, result_label):
    username_to_search = username_entry.get().lower()

    try:
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active

        matching_entries = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 7 and username_to_search in str(row[8]).lower():
                if row[9]=="None":
                    matching_entries.append((row[8], "No tiene permisos"))
                else:
                    matching_entries.append((row[8], row[9]))
                print(username_to_search, "  y  ", str(row[8]).lower())
        if matching_entries:
            result_text = "Usuarios encontrados:\n"
            for entry in matching_entries:
                result_text += f"{entry[0]} {entry[1]}\n"
            result_label.config(text=result_text)
        else:
            result_label.config(text=f"No existe ningún usuario con el nombre '{username_to_search}'.")

    except FileNotFoundError:
        result_label.config(text="Error: Archivo no encontrado.")
    except Exception as e:
        result_label.config(text=f"Error: {str(e)}")

def open_search_window(root):
    search_window = tk.Toplevel(root)
    search_window.title("Buscar Usuario")

    username_entry = tk.Entry(search_window)
    username_entry.grid(row=0, column=1, padx=5, pady=5)

    search_button = tk.Button(search_window, text="Buscar", command=lambda: search_data(root, username_entry, result_label))
    search_button.grid(row=0, column=2, padx=5, pady=5)

    result_label = tk.Label(search_window, text="")
    result_label.grid(row=1, column=0, columnspan=3, padx=5, pady=5)