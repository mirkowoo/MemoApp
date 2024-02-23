from openpyxl import Workbook, load_workbook
from utils import clear_entries, update_table
from openpyxl.styles import PatternFill, Border, Side
import tkinter as tk

def add_data(column_headers, entry_fields, departamentos_combobox, tipo_acceso_combobox, ejecutando_combobox, table_text):
    data = [entry_fields[header].get() if header not in ['Tipo_acceso', 'Ejecutando', 'Depto_solicitante'] else tipo_acceso_combobox.get() if header == 'Tipo_acceso' else ejecutando_combobox.get(
    ) if header == 'Ejecutando' else departamentos_combobox.get() if header == 'Depto_solicitante' else entry_fields[header].get() for header in column_headers]
    # print(f"Departamentos value: {departamentos_combobox.get()}")

    try:
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(column_headers)

        sheet.sheet_view.showGridLines = True

        # Add grid styling
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    sheet.append(data)
    workbook.save('data.xlsx')

    # Update the table display
    update_table(table_text)

    # Clear the entry fields and comboboxes
    clear_entries(entry_fields)
    tipo_acceso_combobox.set("")
    ejecutando_combobox.set("")
    departamentos_combobox.set("")

    table_text.config(state=tk.DISABLED)